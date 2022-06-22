import pandas as pd

pd.options.mode.chained_assignment = None
import argparse
import os
import yaml
from openpyxl import load_workbook
from preprocessing.dataset_preprocessing import preprocess_dataset


def extract_excel(short_names, long_names, excel_file, save_folder):
    # Get only sheet names that exist
    wb = load_workbook(excel_file, read_only=True)
    stocks_exist = [(short_name, long_name) for (short_name, long_name) in zip(short_names, long_names) if
                    short_name in wb.sheetnames]
    short_names_exist = [short_name for (short_name, _) in stocks_exist]
    long_names_exist = [long_name for (_, long_name) in stocks_exist]

    # Iterate over all sheets in the Excel file
    for i_stock, (short_name, long_name) in enumerate(zip(short_names_exist, long_names_exist)):
        long_name_u = long_name.upper()

        # Edit current ae
        dataset = pd.read_excel(excel_file, sheet_name=short_name)[2:]
        dataset = dataset.reset_index(drop=True)
        dataset = dataset.rename(columns={dataset.columns[0]: "date",
                                          f"{long_name_u}": "close",
                                          f"{long_name_u} - PRICE HIGH": "high",
                                          f"{long_name_u} - PRICE LOW": "low",
                                          f"{long_name_u} - OPENING PRICE": "open",
                                          f"{long_name_u} - TURNOVER BY VOLUME": "volume"})

        # Preprocess the ae
        final_dataset = preprocess_dataset(dataset)
        print(f"Completed: {round((i_stock + 1) / len(short_names) * 100, 2)}%")

        # Save the ae
        filepath = os.path.join(save_folder, short_name) + '.csv'
        final_dataset.to_csv(filepath)
        print(f"Saved stock {long_name}")


def parse_arguments():
    dataset_version = 'dataset_v2'
    parser = argparse.ArgumentParser(description='Extract ae from Excel Spreadsheet ae.')
    parser.add_argument('-excel_filepath', '-excel_filepath',
                        default=f'../../Datasets/{dataset_version}/0_raw_data/dataset.xlsx', type=str,
                        help='Filepath to the Excel ae.')
    parser.add_argument('-save_directory', '-save_directory',
                        default=f'../../Datasets/{dataset_version}/1_preprocessed_data', type=str,
                        help='Directory to save the ae.')
    parser.add_argument('-yaml_stocks_filepath', '-yaml_stocks_filepath', default='stock_names.yml', type=str,
                        help='Filepath to the YAML file with the list of stocks')
    args = parser.parse_args()

    if not os.path.exists(args.excel_filepath):
        raise FileNotFoundError('Cannot find Excel file with the ae!')

    return args


def get_stocks(args):
    content = {}
    config_folder = 'config_files'
    with open(os.path.join(config_folder, args.yaml_stocks_filepath)) as f:
        content = yaml.safe_load(f)
    if 'short_names' not in content.keys() or 'long_names' not in content.keys():
        raise FileNotFoundError('The provided YAML file is not well written!')
    if len(content['short_names']) != len(content['long_names']):
        raise FileNotFoundError(
            'The provided YAML file does not have the same number of long stock names and short stock names!')
    return content['short_names'], content['long_names']


def main():
    args = parse_arguments()
    short_names, long_names = get_stocks(args)
    extract_excel(short_names, long_names, args.excel_filepath, args.save_directory)


if __name__ == '__main__':
    main()
